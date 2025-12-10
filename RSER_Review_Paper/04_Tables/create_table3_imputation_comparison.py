"""
Table 3: Imputation Methods Comparison
Comprehensive comparison of imputation strategies used in biomass pyrolysis ML
Data sources: TİK-2, TİK-3, and literature review
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Imputation methods comparison data
data = {
    'Method': [
        'Mean/Median Imputation',
        'Forward/Backward Fill',
        'Deletion (Listwise)',
        'Deletion (Pairwise)',
        'K-Nearest Neighbors (KNN)',
        'Multiple Imputation (MICE)',
        'Iterative Imputer',
        'Random Forest (MissForest)',
        'Domain Knowledge-Based',
        'GAN-based Synthetic Data'
    ],
    'Complexity': [
        'Very Low',
        'Very Low',
        'N/A',
        'N/A',
        'Medium',
        'High',
        'Medium',
        'High',
        'Low-Medium',
        'Very High'
    ],
    'Computational_Cost': [
        'Very Low',
        'Very Low',
        'N/A',
        'N/A',
        'Low-Medium',
        'High',
        'Medium',
        'High',
        'Very Low',
        'Very High'
    ],
    'Preserves_Relationships': [
        'No',
        'Partial',
        'Yes (removes data)',
        'Yes (removes data)',
        'Yes',
        'Yes',
        'Yes',
        'Yes (best)',
        'Yes (if designed properly)',
        'Yes (learned)'
    ],
    'Best_Use_Case': [
        'Low variance variables (e.g., Nitrogen)',
        'Time series data (not applicable here)',
        'Very low missing % (<5%)',
        'Specific analysis tasks only',
        'Correlated features (Volatiles, HHV)',
        'Multiple correlated outputs',
        'Mixed data types',
        'Complex non-linear relationships',
        'Variables with known formulas (O/C, H/C, Holocellulose)',
        'Extremely small datasets needing augmentation'
    ],
    'Advantages': [
        'Simple; fast; no tuning needed',
        'Preserves temporal order',
        'No imputation bias; clean data',
        'Maximizes sample usage per analysis',
        'Preserves local patterns; moderate accuracy',
        'Uncertainty quantification; robust',
        'Flexible; handles mixed types',
        'High accuracy; preserves correlations',
        'Exact (no estimation error); chemically consistent',
        'Generates new plausible samples; addresses overfitting'
    ],
    'Disadvantages': [
        'Distorts variance; ignores correlations',
        'Not applicable for cross-sectional data',
        'Reduces sample size; potential bias',
        'Different N for each analysis; hard to reproduce',
        'Sensitive to k choice; struggles with categorical',
        'Computationally expensive; complex to implement',
        'Can be unstable; requires careful tuning',
        'Slow for large datasets; hyperparameter tuning needed',
        'Requires domain expertise; limited applicability',
        'Very high computational cost; risk of mode collapse'
    ],
    'Used_in_TİK_Reports': [
        'Yes (Nitrogen)',
        'No',
        'Yes (segregated dataset strategy)',
        'No',
        'Yes (Volatiles, FixedCarbon, HHV, Cellulose, Hemicellulose)',
        'No',
        'No',
        'No (but mentioned in literature)',
        'Yes (O/C, H/C, Holocellulose, Duration)',
        'No (but literature Ref [12] used)'
    ],
    'Typical_Performance': [
        'Baseline',
        'N/A',
        'N/A',
        'N/A',
        'Moderate (R²≈0.80 in lit)',
        'Good (R²≈0.85)',
        'Good (R²≈0.85)',
        'Best (R²≈0.90-0.95 in lit)',
        'Exact (if formula-based)',
        'Excellent (88.98% accuracy in Ref [12])'
    ],
    'Recommended_Missing_%': [
        '<20%',
        'N/A',
        '<5%',
        '<10%',
        '10-50%',
        '10-50%',
        '10-50%',
        '10-70%',
        'Any (if formula exists)',
        '>50% (augmentation)'
    ],
    'Python_Library': [
        'numpy.nanmean / pandas.fillna',
        'pandas.fillna (method="ffill"/"bfill")',
        'pandas.dropna (how="any")',
        'Custom per analysis',
        'sklearn.impute.KNNImputer',
        'sklearn.experimental.IterativeImputer / statsmodels',
        'sklearn.impute.IterativeImputer',
        'missingpy.MissForest / missForest (R)',
        'Custom functions',
        'TensorFlow / PyTorch (custom GAN)'
    ]
}

df = pd.DataFrame(data)

# Create Excel file
output_path = 'C:\\@biyokomurlestirme\\RSER_Review_Paper\\04_Tables\\Table3_Imputation_Comparison.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Write main comparison table
    df.to_excel(writer, sheet_name='Imputation_Comparison', index=False)

    workbook = writer.book
    worksheet = writer.sheets['Imputation_Comparison']

    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    used_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    not_used_fill = PatternFill(start_color='FFCCBC', end_color='FFCCBC', fill_type='solid')

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Format header
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
        used_in_tik = df.iloc[row_idx-2]['Used_in_TİK_Reports']

        # Highlight if used in TİK reports
        if 'Yes' in used_in_tik:
            row[7].fill = used_fill  # Used_in_TİK_Reports column
        else:
            row[7].fill = not_used_fill

        # Apply borders and alignment
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Adjust column widths
    worksheet.column_dimensions['A'].width = 28
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 18
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 35
    worksheet.column_dimensions['F'].width = 40
    worksheet.column_dimensions['G'].width = 40
    worksheet.column_dimensions['H'].width = 35
    worksheet.column_dimensions['I'].width = 25
    worksheet.column_dimensions['J'].width = 20
    worksheet.column_dimensions['K'].width = 45

    # Freeze header row
    worksheet.freeze_panes = 'A2'

    # Create recommendations sheet
    recommendations = {
        'Scenario': [
            'Very low missing data (<5%)',
            'Low missing data (5-20%)',
            'Moderate missing data (20-50%)',
            'High missing data (50-80%)',
            'Critical missing data (>80%)',
            '',
            'Variables with known formulas',
            'Correlated biomass characterization',
            'Process parameters (temporal)',
            'Output variables (targets)',
            '',
            'Small dataset (N<50)',
            'Medium dataset (50<N<200)',
            'Large dataset (N>200)',
            '',
            'When speed is critical',
            'When accuracy is critical',
            'When interpretability is critical'
        ],
        'Recommended_Method': [
            'Deletion (listwise)',
            'Mean imputation',
            'KNN imputation (k=5)',
            'MissForest or KNN',
            'Domain knowledge + GAN augmentation',
            '',
            'Calculation-based (ALWAYS prefer)',
            'KNN (k=3-5)',
            'Mean or forward fill',
            'Per-model cleanup (segregated dataset)',
            '',
            'Simple methods (mean, KNN) to avoid overfitting',
            'KNN or Iterative Imputer',
            'MissForest or MICE',
            '',
            'Mean imputation',
            'MissForest or domain knowledge',
            'Mean or domain knowledge'
        ],
        'Rationale': [
            'Minimal data loss; clean interpretation',
            'Low variance disruption; fast',
            'Balance of accuracy and speed; preserves local patterns',
            'Need sophisticated methods to preserve relationships',
            'Cannot impute reliably; need data augmentation',
            '',
            'Zero estimation error; chemically consistent',
            'Preserves inter-variable correlations in biomass composition',
            'Temporal dependence if batch vs continuous',
            'Keep missing to train separate models per output',
            '',
            'Avoid overfitting from complex imputation',
            'Can handle moderate complexity without overfitting',
            'Enough data to support complex methods',
            '',
            'Fastest computation',
            'Best preservation of variance and relationships',
            'Transparent logic; easy to explain'
        ]
    }

    rec_df = pd.DataFrame(recommendations)
    rec_df.to_excel(writer, sheet_name='Recommendations', index=False)

    rec_ws = writer.sheets['Recommendations']

    # Format recommendations header
    for cell in rec_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    rec_ws.column_dimensions['A'].width = 40
    rec_ws.column_dimensions['B'].width = 45
    rec_ws.column_dimensions['C'].width = 60

    # Bold section headers (empty rows before sections)
    for i in [7, 12, 16]:
        if i <= len(rec_df):
            rec_ws[f'A{i}'].font = Font(bold=True, size=11)

    # Create TİK-specific strategy sheet
    tik_strategy = {
        'Step': [
            '1',
            '2',
            '3',
            '4',
            '5',
            '6'
        ],
        'Variable_Group': [
            'Elemental Ratios',
            'Structural Components',
            'Biomass Properties',
            'Process Parameters',
            'Bio-oil Outputs',
            'Final Validation'
        ],
        'Variables': [
            'O/C, H/C',
            'Holocellulose, Cellulose, Hemicellulose',
            'Volatiles, FixedCarbon, HHV',
            'Duration (unified from FeedRate + ResidenceTime), GasFlowrate',
            'All 11 chemical groups',
            'Check constraints and physical consistency'
        ],
        'Method_Used': [
            'Calculation-based',
            'Calculation + KNN with constraint',
            'KNN (k=5)',
            'Formula synthesis + Mean',
            'Segregated dataset (per-output modeling)',
            'Manual inspection + domain rules'
        ],
        'Reason': [
            'Exact formula: O/C = O% / C%',
            'Holocellulose = Cellulose + Hemicellulose; scale to match',
            'High correlation with elemental composition (O/C, H/C, S, Ash)',
            'FeedRate + ResidenceTime → Duration; GasFlowrate low impact',
            'High missing % for some; separate models avoid NULL contamination',
            'Ensure C+H+O+N+S+Ash ≈ 100%, bio-oil components sum ≤ 100%'
        ],
        'Success_Metric': [
            '0% missing after calculation',
            '~15% missing after KNN',
            '~5% missing after KNN',
            '~8% missing after synthesis',
            'Each model trained on complete cases only',
            'All constraints satisfied'
        ]
    }

    tik_df = pd.DataFrame(tik_strategy)
    tik_df.to_excel(writer, sheet_name='TİK_Strategy', index=False)

    tik_ws = writer.sheets['TİK_Strategy']

    # Format TİK strategy header
    for cell in tik_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Adjust column widths
    tik_ws.column_dimensions['A'].width = 8
    tik_ws.column_dimensions['B'].width = 25
    tik_ws.column_dimensions['C'].width = 50
    tik_ws.column_dimensions['D'].width = 40
    tik_ws.column_dimensions['E'].width = 60
    tik_ws.column_dimensions['F'].width = 45

    # Format cells
    for row in tik_ws.iter_rows(min_row=2, max_row=len(tik_df)+1):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

print(f"Table 3 created successfully: {output_path}")
print("\n=== IMPUTATION METHODS SUMMARY ===")
print(f"Total methods compared: {len(df)}")
print(f"Methods used in TİK reports: {len(df[df['Used_in_TİK_Reports'].str.contains('Yes')])}")
print(f"\nBest performers:")
print(f"  - Random Forest (MissForest): R²≈0.90-0.95")
print(f"  - Domain Knowledge-Based: Exact (0% error)")
print(f"  - KNN: Moderate (R²≈0.80)")
print(f"\nTİK Reports Strategy: Multi-tier approach")
print(f"  1. Calculation-based (O/C, H/C, Holocellulose, Duration)")
print(f"  2. KNN imputation (Volatiles, HHV, Cellulose)")
print(f"  3. Mean imputation (Nitrogen, GasFlowrate)")
print(f"  4. Segregated dataset for outputs")
