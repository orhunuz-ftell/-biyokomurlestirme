"""
Table 2: Missing Data Analysis
Creates comprehensive Excel table showing missing data patterns
Data source: TİK-2 Report
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Missing data statistics from TİK-2
data = {
    'Variable': [
        # Biomass Characterization
        'Carbon (C)', 'Hydrogen (H)', 'Nitrogen (N)', 'Oxygen (O)',
        'Sulfur (S)', 'HHV', 'Volatiles', 'FixedCarbon', 'Ash',
        # Structural Components
        'Cellulose', 'Hemicellulose', 'Lignin', 'Holocellulose',
        # Process Parameters
        'ProcessTemperature', 'CatalystBiomassRatio', 'FeedRate',
        'ResidenceTime', 'GasFlowrate', 'Duration',
        # Bio-oil Outputs
        'LiquidOutput', 'Sugar', 'Alcohols', 'Aromatics', 'Furans',
        'Aldehyde_ketone', 'Acids', 'Phenols', 'Oxides', 'Esters',
        'Aliphatichydrocarbon'
    ],
    'Category': [
        # Biomass
        'Biomass Characterization', 'Biomass Characterization', 'Biomass Characterization', 'Biomass Characterization',
        'Biomass Characterization', 'Biomass Characterization', 'Biomass Characterization', 'Biomass Characterization', 'Biomass Characterization',
        # Structural
        'Structural Components', 'Structural Components', 'Structural Components', 'Structural Components',
        # Process
        'Process Parameters', 'Process Parameters', 'Process Parameters',
        'Process Parameters', 'Process Parameters', 'Process Parameters',
        # Outputs
        'Bio-oil Composition', 'Bio-oil Composition', 'Bio-oil Composition', 'Bio-oil Composition', 'Bio-oil Composition',
        'Bio-oil Composition', 'Bio-oil Composition', 'Bio-oil Composition', 'Bio-oil Composition', 'Bio-oil Composition',
        'Bio-oil Composition'
    ],
    'Total_Count': [48] * 30,  # All from 48 samples
    'Missing_Count': [
        # Biomass
        0, 0, 0, 0, 15, 10, 5, 5, 8,
        # Structural
        18, 18, 12, 18,
        # Process
        0, 0, 43, 43, 23, 8,
        # Outputs
        18, 27, 25, 14, 11, 5, 2, 8, 25, 23, 23
    ],
    'Missing_%': [
        # Biomass
        0.00, 0.00, 0.00, 0.00, 31.25, 20.83, 10.42, 10.42, 16.67,
        # Structural
        37.50, 37.50, 25.00, 37.50,
        # Process
        0.00, 0.00, 89.58, 89.58, 47.92, 16.67,
        # Outputs
        37.50, 56.25, 52.08, 29.17, 22.92, 10.42, 4.17, 16.67, 52.08, 47.92, 47.92
    ],
    'Priority': [
        # Biomass
        'Complete', 'Complete', 'Complete', 'Complete', 'Medium', 'Low', 'Low', 'Low', 'Low',
        # Structural
        'Medium', 'Medium', 'Medium', 'Medium',
        # Process
        'Complete', 'Complete', 'CRITICAL', 'CRITICAL', 'High', 'Low',
        # Outputs
        'Medium', 'High', 'High', 'Medium', 'Medium', 'Low', 'Complete', 'Low', 'High', 'High', 'High'
    ],
    'Imputation_Method': [
        # Biomass
        'None', 'None', 'None', 'None', 'Mean', 'KNN', 'KNN', 'KNN', 'Literature lookup',
        # Structural
        'KNN', 'KNN', 'KNN', 'Calculation',
        # Process
        'None', 'None', 'Duration synthesis', 'Duration synthesis', 'Mean', 'Calculation',
        # Outputs
        'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup',
        'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup', 'Per-model cleanup'
    ],
    'Impact_on_Model': [
        # Biomass
        'High (core feature)', 'High (core feature)', 'High (core feature)', 'High (core feature)',
        'Medium', 'Medium', 'Medium', 'Medium', 'Low',
        # Structural
        'High', 'High', 'High', 'Medium (derived)',
        # Process
        'Critical', 'Medium', 'Critical (synthesized)', 'Critical (synthesized)', 'High', 'High',
        # Outputs (these are targets, not features)
        'Target variable', 'Target variable', 'Target variable', 'Target variable', 'Target variable',
        'Target variable', 'Target variable', 'Target variable', 'Target variable', 'Target variable', 'Target variable'
    ]
}

df = pd.DataFrame(data)

# Sort by Missing % (descending)
df = df.sort_values('Missing_%', ascending=False)

# Create Excel file with formatting
output_path = 'C:\\@biyokomurlestirme\\RSER_Review_Paper\\04_Tables\\Table2_MissingData_Analysis.xlsx'

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Write main table
    df.to_excel(writer, sheet_name='Missing_Data_Summary', index=False)

    workbook = writer.book
    worksheet = writer.sheets['Missing_Data_Summary']

    # Define styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    critical_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    high_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    medium_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    low_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    complete_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Format data rows
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
        missing_pct = df.iloc[row_idx-2]['Missing_%']
        priority = df.iloc[row_idx-2]['Priority']

        # Color code based on priority
        if priority == 'CRITICAL':
            fill = critical_fill
        elif priority == 'High':
            fill = high_fill
        elif priority == 'Medium':
            fill = medium_fill
        elif priority == 'Low':
            fill = low_fill
        else:  # Complete
            fill = complete_fill

        # Apply formatting to priority cell
        row[4].fill = fill

        # Apply borders to all cells
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Center align numeric columns
        row[2].alignment = Alignment(horizontal='center', vertical='center')
        row[3].alignment = Alignment(horizontal='center', vertical='center')
        row[4].alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 22
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 18
    worksheet.column_dimensions['G'].width = 25
    worksheet.column_dimensions['H'].width = 22

    # Freeze header row
    worksheet.freeze_panes = 'A2'

    # Create summary sheet
    summary_data = {
        'Metric': [
            'Total Variables Analyzed',
            'Variables with 0% Missing',
            'Variables with >50% Missing',
            'Variables with >80% Missing (CRITICAL)',
            '',
            'Average Missing % (All)',
            'Average Missing % (Biomass Characterization)',
            'Average Missing % (Structural Components)',
            'Average Missing % (Process Parameters)',
            'Average Missing % (Bio-oil Composition)',
            '',
            'Most Complete Variable',
            'Most Incomplete Variable',
            '',
            'Imputation Methods Used',
            '  - Calculation-based',
            '  - KNN imputation',
            '  - Mean imputation',
            '  - Per-model cleanup'
        ],
        'Value': [
            len(df),
            len(df[df['Missing_%'] == 0]),
            len(df[df['Missing_%'] > 50]),
            len(df[df['Missing_%'] > 80]),
            '',
            f"{df['Missing_%'].mean():.2f}%",
            f"{df[df['Category']=='Biomass Characterization']['Missing_%'].mean():.2f}%",
            f"{df[df['Category']=='Structural Components']['Missing_%'].mean():.2f}%",
            f"{df[df['Category']=='Process Parameters']['Missing_%'].mean():.2f}%",
            f"{df[df['Category']=='Bio-oil Composition']['Missing_%'].mean():.2f}%",
            '',
            f"{df.iloc[-1]['Variable']} (0%)",
            f"{df.iloc[0]['Variable']} ({df.iloc[0]['Missing_%']:.2f}%)",
            '',
            'Multiple strategies employed',
            'O/C, H/C, Holocellulose, Duration',
            'Volatiles, FixedCarbon, HHV, Cellulose, Hemicellulose',
            'Sulfur, GasFlowrate',
            'Bio-oil composition outputs'
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary_Statistics', index=False)

    summary_ws = writer.sheets['Summary_Statistics']

    # Format summary sheet header
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 45
    summary_ws.column_dimensions['B'].width = 40

    # Bold section headers
    for row in [6, 12, 15]:
        summary_ws[f'A{row}'].font = Font(bold=True, size=11)

print(f"Table 2 created successfully: {output_path}")

# Print summary
print("\n=== MISSING DATA ANALYSIS SUMMARY ===")
print(f"Total variables: {len(df)}")
print(f"Complete (0% missing): {len(df[df['Missing_%'] == 0])}")
print(f"Critical (>80% missing): {len(df[df['Missing_%'] > 80])}")
print(f"High priority (50-80% missing): {len(df[(df['Missing_%'] > 50) & (df['Missing_%'] <= 80)])}")
print(f"\nAverage missing data: {df['Missing_%'].mean():.2f}%")
print(f"\nTop 5 most incomplete variables:")
for i, row in df.head(5).iterrows():
    print(f"  • {row['Variable']}: {row['Missing_%']:.2f}% missing")
